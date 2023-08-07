from openpyxl import load_workbook
from openpyxl.styles import PatternFill


pathRead = "D:\Reps\WritingInExcel\WritingInExcel\ReadBooks.xlsx"
readBooks = load_workbook(pathRead)
sheet = readBooks.active

pathToRead = "D:\Reps\WritingInExcel\WritingInExcel\ToReadBooks.xlsx"
toReadBooks = load_workbook(pathToRead)
sheetToRead = toReadBooks.active


# Insert a row into the worksheet at 3rd position


# Save the modified Excel file in default (that is Excel 2003) format

book =''
author=''
language =''
startingDate =''
finishedDate =''
description = ""
linkToMedium = ""

class AddingBook():


        
    
    def setStartingDate(self, startingDate2):
        global startingDate
        startingDate = startingDate2
       


    def setFinishedDate(self,finishedDate2):
        global finishedDate
        finishedDate = finishedDate2
        


    def setAuthor(self, author2):
        global author
        author = author2
        
    
    def setLanguage(self,language2):
        global language
        language = language2
   
    
    def setBookName(self,book2):
        global book
        book = book2

    def setDescription(self, description2):
        global description
        description = description2
    
    def setLinkToMedium(self, link):
        global linkToMedium
        linkToMedium = link


    def insertRead(self):
        sheet.insert_rows(idx=2)
        bookName =  sheet['A2']
        bookName.value = book
        authorName =  sheet['B2']
        authorName.value = author
        languageType =  sheet['C2']
        languageType.value = language
        startingDt =  sheet['D2']
        startingDt.value = startingDate
        finishedDt =  sheet['E2']
        finishedDt.value = finishedDate
        descriptionCell = sheet['F2']
        descriptionCell.value = description
        linkToMediumCell = sheet['G2']
        linkToMediumCell.value = linkToMedium
        readBooks.save(pathRead)

    def insertToRead(self):
        sheetToRead.insert_rows(idx= 2)
        for rows in sheetToRead.iter_rows(min_row=2, max_row=3, min_col=1, max_col=3):
            for cells in rows:
                yellow = "00FFFF00"
                cells.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")

        bookName = sheetToRead['A2']
        bookName.value = book
        authorName = sheetToRead['B2']
        authorName.value = author
        languageType = sheetToRead['C2']
        languageType.value = language
        toReadBooks.save(pathToRead)
    


  
        
    

    
